import re
import openpyxl
from pygoogle_image import image as pi

# Function to clean the data
def clean_data(data):
    return re.sub(r'[^A-Za-z0-9 ]', '', data)

# Load data from Excel file
input_file = 'input_data.xlsx'
output_file = 'image_urls.xlsx'

wb = openpyxl.load_workbook(input_file)
sheet = wb.active

# Open Excel file for writing image URLs
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.append(['title', 'image_url'])  # Write header row

for row in sheet.iter_rows(values_only=True):
    if row:
        title = row[0]

        # Clean the data
        clean_title = clean_data(title)
        print("Cleaned Title:", clean_title)

        try:
            # Download image based on the cleaned data
            urls = pi.download(clean_title, limit=1)
            # urls = "/images/"
            print("Image URLs:", urls[1])

            if urls:

                # Write original title, image URL, and file name to Excel
                ws_output.append([title, urls[1]])
        except Exception as e:
            print(f"Error occurred for '{title}': {e}")

# Adjust column width
for col in ws_output.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws_output.column_dimensions[column].width = adjusted_width

# Save the output Excel file
wb_output.save(output_file)

print("Image URLs and file names have been written to 'image_urls.xlsx'.")
